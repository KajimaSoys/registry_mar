from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from . import data_format


def waybill_generate(data):
    obj = data_format.format(data)
    print(obj)

    sourcePath = 'static/xlsx/check-template.xlsx'
    filename = obj['path']
    exitPath = f'media/waybills/{filename}.xlsx'
    # exitPath = f'rezal/apps/main/contract_db/{obj["path"]}/{obj["path"]}.xlsx'

    wb = load_workbook(sourcePath)
    sheet = wb.get_sheet_by_name('Waybill')
    wb.active = sheet
    ws = wb.active

    transfer_act(ws, obj)

    wb.save(exitPath)
    return exitPath

def transfer_act(ws, obj):
    border_all = Border(left=Side(border_style='thin', color = '000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    border_without_right = Border(left=Side(border_style='thin', color = '000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    border_without_left = Border(right=Side(border_style='thin', color='000000'),
                                  top=Side(border_style='thin', color='000000'),
                                  bottom=Side(border_style='thin', color='000000'))


    ws['A2'] = f'Акт оказанных услуг № {obj["id"]} от «{obj["date_start_word"][0]}» {obj["date_start_word"][1]} {obj["date_start_word"][2]} года'

    ws['A6'] = f'ООО "Ахмадуллина", именуемое в дальнейшем "Исполнитель", и пациент {obj["patient_name"]} {obj["born"]} г/р' \
               f', проживающий (-ая) по адресу {obj["address"]}, именуемый (-ая) в дальнейшем ' \
               '"Заказчик", настоящим актом подтверждают, что Заказчику были оказаны следующие медицинские услуги: '

    ws['A10'] = f'Лечащий врач: {obj["doctor_name"]}'
    ws['A11'] = f'Заказчик: {obj["patient_name"]}'

    last_row = '16'
    i = 0
    try:
        for row in ws['A15': 'N25']:
            for cell in row:
                obj['service_name'][i]
                if cell.column == 1:
                    ws[cell.coordinate] = i + 1
                if cell.column == 2:
                    # ws.merge_cells(start_row=cell.row, start_column=2, end_row=cell.row, end_column=21)
                    ws[cell.coordinate] = obj['service_name'][i]
                if cell.column == 10:
                    ws[cell.coordinate] = obj['cost'][i]
                if cell.column == 12:
                    ws[cell.coordinate] = 1
                if cell.column == 13:
                    ws[cell.coordinate] = f'=J{cell.row}*L{cell.row}'
                ws[cell.coordinate].border = border_all
            i += 1
    except Exception:
        last_row = cell.row
        print(last_row)

    for row in ws[f'A{last_row}':f'N{last_row}']:
        for cell in row:
            if cell.column == 1:
                ws[cell.coordinate].border = border_without_right
            if cell.column == 2:
                ws.unmerge_cells(start_row=cell.row, end_row=cell.row, start_column=2, end_column=9)
                ws.unmerge_cells(start_row=cell.row, end_row=cell.row, start_column=10, end_column=11)
                ws.merge_cells(start_row=cell.row, start_column=2, end_row=cell.row, end_column=12)

                ws[cell.coordinate] = "Итого:"
                ws[cell.coordinate].border = border_without_left
                ws[cell.coordinate].font = Font(name='Times New Roman', b=True)
                ws[cell.coordinate].alignment = Alignment(horizontal="left", vertical="center")
            if cell.column > 1 and cell.column < 15:
                ws[cell.coordinate].border = border_all
                # ws[cell.coordinate].border = Border(bottom=Side(border_style='thin'))
            if cell.column == 13:
                ws[cell.coordinate] = f'=SUM(M15:M{cell.row-1})'
                # ws[cell.coordinate].border = border_all
    last_row = cell.row + 1
    for row in ws[f'A{last_row}':f'N{last_row}']:
        for cell in row:
            if cell.column == 1:
                ws[cell.coordinate].border = border_without_right
            if cell.column == 2:
                ws.unmerge_cells(start_row=cell.row, end_row=cell.row, start_column=2, end_column=9)
                ws.unmerge_cells(start_row=cell.row, end_row=cell.row, start_column=10, end_column=11)
                ws.merge_cells(start_row=cell.row, start_column=2, end_row=cell.row, end_column=12)

                ws[cell.coordinate] = "К оплате с учетом скидок:"
                ws[cell.coordinate].border = border_without_left
                ws[cell.coordinate].font = Font(name='Times New Roman', b=True)
                ws[cell.coordinate].alignment = Alignment(horizontal="left", vertical="center")
            if cell.column > 1 and cell.column < 15:
                ws[cell.coordinate].border = border_all
                # ws[cell.coordinate].border = Border(bottom=Side(border_style='thin'))
            if cell.column == 13:
                ws[cell.coordinate] = f'{obj["summa"]}'


    ws['A29'] = f"Исполнитель:" + " " * 109 + "Заказчик:\n\n\n\n" \
                "__________________/Ахмадуллина А. Р./" + " " * 60 + f"__________________/{obj['patient_name_output']}/\n" \
                + " " * 18 + "М.П."